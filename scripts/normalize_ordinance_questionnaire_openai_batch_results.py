#!/usr/bin/env python3
"""
Normalize + validate OpenAI batch results for the ordinance questionnaire step.

This repo intentionally does NOT enforce JSON via OpenAI-side schema for the questionnaire step.
Instead, we:
  - ask for a schema in the prompt
  - download raw batch outputs
  - normalize/repair small deviations locally (deterministic)
  - flag truly broken rows for re-run

Inputs:
  - A questionnaire request dir (exported by scripts/export_ordinance_questionnaire_batch_requests.py)
      * mapping_shard*.jsonl  (custom_id -> page_id/box_id/etc)
  - A questionnaire results dir (downloaded by scripts/download_openai_batch_results.py)
      * openai_results_shard*.jsonl
      * openai_errors_shard*.jsonl (optional; may be empty)
  - Questions.xlsx (Processed Info + Input Info) for expected IDs + categorical options

Outputs:
  - normalized.jsonl (one row per custom_id with normalized answers_by_id)
  - invalid_rows.jsonl (only rows that could not be parsed/normalized)
  - summary printed to stdout
"""

from __future__ import annotations

import argparse
import json
import re
import statistics
from collections import Counter
from dataclasses import dataclass
from os.path import expanduser
from pathlib import Path
from typing import Any


def _eprint(msg: str) -> None:
    print(msg, flush=True)


def _load_env_file(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        key = k.strip()
        val = v.strip().strip('"').strip("'")
        if key:
            env[key] = val
    return env


def _coerce_json(text: str) -> dict[str, Any]:
    """Accept raw JSON or JSON wrapped in ```json fences, scrub control chars.

    Also tries a '{...}' substring extraction as a last resort.
    """

    if text is None:
        raise ValueError("empty response text")
    stripped = str(text).strip()
    if stripped.startswith("```"):
        lines = stripped.splitlines()
        lines = lines[1:]
        if lines and lines[-1].strip().startswith("```"):
            lines = lines[:-1]
        stripped = "\n".join(lines).strip()
    try:
        return json.loads(stripped)
    except json.JSONDecodeError:
        cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", stripped)
        try:
            return json.loads(cleaned)
        except json.JSONDecodeError:
            # Some providers return raw newlines in strings; try escaping them and retry.
            cleaned2 = cleaned.replace("\n", "\\n")
            try:
                return json.loads(cleaned2)
            except json.JSONDecodeError:
                # Last resort: take substring between first '{' and last '}'.
                a = cleaned2.find("{")
                b = cleaned2.rfind("}")
                if a >= 0 and b > a:
                    return json.loads(cleaned2[a : b + 1])
                raise


def _extract_openai_output_text(body: dict[str, Any]) -> str:
    out = body.get("output")
    if not isinstance(out, list):
        return ""
    parts: list[str] = []
    for item in out:
        if not isinstance(item, dict):
            continue
        content = item.get("content") or []
        if not isinstance(content, list):
            continue
        for c in content:
            if isinstance(c, dict) and c.get("type") == "output_text":
                parts.append(str(c.get("text") or ""))
    return "".join(parts)


@dataclass(frozen=True)
class Question:
    id: str
    question_type: str
    possible_answers: list[str] | None


def _load_questions_xlsx(*, xlsx_path: Path, processed_sheet: str, input_info_sheet: str) -> list[Question]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise SystemExit("Missing dependency: openpyxl (pip install openpyxl)") from exc

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if processed_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {processed_sheet!r} (have {wb.sheetnames})")
    if input_info_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {input_info_sheet!r} (have {wb.sheetnames})")

    ws_proc = wb[processed_sheet]
    ws_in = wb[input_info_sheet]

    def _hdr_map(ws) -> dict[str, int]:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        out: dict[str, int] = {}
        for i, h in enumerate(headers, start=1):
            if h is None:
                continue
            key = str(h).strip()
            if key:
                out[key] = i
        return out

    proc_idx = _hdr_map(ws_proc)
    in_idx = _hdr_map(ws_in)

    required_proc = {"ID", "Include", "Question Type"}
    missing_proc = sorted(required_proc - set(proc_idx))
    if missing_proc:
        raise SystemExit(f"Processed sheet missing headers: {missing_proc}")

    possible_by_id: dict[str, list[str]] = {}
    if "ID" in in_idx and "Possible Answers" in in_idx:
        for r in range(2, ws_in.max_row + 1):
            qid_raw = ws_in.cell(r, in_idx["ID"]).value
            if qid_raw is None:
                continue
            qid = str(qid_raw).strip()
            if not qid:
                continue
            poss_raw = ws_in.cell(r, in_idx["Possible Answers"]).value
            if poss_raw is None:
                continue
            poss_s = str(poss_raw).strip()
            if not poss_s:
                continue
            opts = [p.strip() for p in poss_s.split(";") if p.strip()]
            if opts:
                possible_by_id[qid] = opts

    questions: list[Question] = []
    for r in range(2, ws_proc.max_row + 1):
        include_raw = ws_proc.cell(r, proc_idx["Include"]).value
        if str(include_raw).strip().lower() != "yes":
            continue
        qid_raw = ws_proc.cell(r, proc_idx["ID"]).value
        if qid_raw is None:
            continue
        qid = str(qid_raw).strip()
        qtype_raw = ws_proc.cell(r, proc_idx["Question Type"]).value
        qtype = str(qtype_raw).strip() if qtype_raw is not None else ""
        if qtype not in {"Binary", "Categorical", "Numerical", "Continuous"}:
            raise SystemExit(f"Unsupported Question Type {qtype!r} for ID={qid} (row {r})")
        questions.append(Question(id=qid, question_type=qtype, possible_answers=possible_by_id.get(qid)))

    if not questions:
        raise SystemExit("No included questions found (Include == Yes)")
    return questions


def _read_mapping(request_dir: Path) -> dict[str, dict[str, Any]]:
    mapping_by_id: dict[str, dict[str, Any]] = {}
    mapping_paths = sorted(request_dir.glob("mapping_shard*.jsonl"))
    if not mapping_paths:
        raise SystemExit(f"No mapping_shard*.jsonl found in {request_dir}")
    for mp in mapping_paths:
        for raw in mp.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError:
                continue
            key = obj.get("id") or obj.get("custom_id")
            if isinstance(key, str) and key:
                mapping_by_id[key] = obj
    return mapping_by_id


def _parse_answers_object(parsed: dict[str, Any]) -> tuple[dict[str, Any] | None, list[str]]:
    """Return (answers_by_id, issues). Does not coerce to expected question set."""
    issues: list[str] = []
    if not isinstance(parsed, dict):
        return None, ["not_object"]

    if isinstance(parsed.get("answers_by_id"), dict):
        return parsed["answers_by_id"], issues

    # Back-compat: some outputs may still use list-of-answers form.
    answers_list = parsed.get("answers")
    if isinstance(answers_list, list):
        out: dict[str, Any] = {}
        for item in answers_list:
            if not isinstance(item, dict):
                continue
            qid = item.get("id")
            if not isinstance(qid, str) or not qid:
                continue
            out[qid] = {
                "answer": item.get("answer"),
                "unit": item.get("unit"),
                "evidence": item.get("evidence"),
                "confidence": item.get("confidence"),
            }
        issues.append("converted_from_answers_list")
        return out, issues

    return None, ["missing_answers"]


def _as_float(v: Any) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        try:
            return float(s)
        except Exception:
            return None
    return None


def _coerce_binary(v: Any) -> bool | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)) and v in {0, 1}:
        return bool(int(v))
    if isinstance(v, str):
        s = v.strip().lower()
        if s in {"true", "t", "yes", "y"}:
            return True
        if s in {"false", "f", "no", "n"}:
            return False
    return None


def _coerce_number(v: Any) -> float | int | None:
    if v is None:
        return None
    if isinstance(v, bool):
        # Avoid interpreting True/False as 1/0 for numeric questions.
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # Conservative parse: only accept pure numbers.
        if not re.fullmatch(r"[-+]?\\d+(?:\\.\\d+)?", s):
            return None
        if "." in s:
            try:
                return float(s)
            except Exception:
                return None
        try:
            return int(s)
        except Exception:
            return None
    return None


def _normalize_one(
    *,
    custom_id: str,
    parsed: dict[str, Any],
    questions: list[Question],
) -> tuple[dict[str, Any] | None, list[str]]:
    """Normalize into the canonical prompt schema: {answers_by_id: {...}, notes: str}."""

    issues: list[str] = []
    answers_raw, issues2 = _parse_answers_object(parsed)
    issues.extend(issues2)
    if answers_raw is None:
        return None, issues

    expected_ids = [q.id for q in questions]
    q_by_id = {q.id: q for q in questions}

    # Drop extras and add missing as nulls.
    got_keys = set(str(k) for k in answers_raw.keys())
    expected_set = set(expected_ids)
    missing = sorted(expected_set - got_keys)
    extra = sorted(got_keys - expected_set)
    if missing:
        issues.append(f"filled_missing_ids:{len(missing)}")
    if extra:
        issues.append(f"dropped_extra_ids:{len(extra)}")

    answers_norm: dict[str, Any] = {}
    for qid in expected_ids:
        q = q_by_id[qid]
        raw = answers_raw.get(qid) if isinstance(answers_raw, dict) else None
        if not isinstance(raw, dict):
            raw = {}

        ans = raw.get("answer", None)
        unit = raw.get("unit", None)
        evidence = raw.get("evidence", None)
        conf = _as_float(raw.get("confidence", None))

        if q.question_type == "Binary":
            ans2 = _coerce_binary(ans)
        elif q.question_type in {"Numerical", "Continuous"}:
            ans2 = _coerce_number(ans)
        elif q.question_type == "Categorical":
            if isinstance(ans, str):
                ans2 = ans.strip()
                if ans2 == "":
                    ans2 = None
            else:
                ans2 = None
            if q.possible_answers and isinstance(ans2, str) and ans2 not in q.possible_answers:
                issues.append(f"{qid}:categorical_not_in_options")
                ans2 = None
        else:
            issues.append(f"{qid}:unknown_type")
            ans2 = None

        # Normalize unit/evidence
        if isinstance(unit, str) and unit.strip() == "":
            unit = None
        if isinstance(evidence, str) and evidence.strip() == "":
            evidence = None

        # Normalize confidence
        if conf is None:
            conf = 0.0
        if conf < 0.0:
            conf = 0.0
        if conf > 1.0:
            conf = 1.0

        if ans2 is None:
            # Prompt rule: null => evidence/unit null + confidence 0.
            if conf != 0.0:
                issues.append(f"{qid}:forced_confidence_zero_for_null")
            ans2 = None
            unit = None
            evidence = None
            conf = 0.0
        else:
            # Evidence is ideal but not mandatory for normalization; flag if missing.
            if evidence is None:
                issues.append(f"{qid}:missing_evidence_for_nonnull")

            # Unit must be null for Binary/Categorical.
            if q.question_type in {"Binary", "Categorical"} and unit is not None:
                issues.append(f"{qid}:dropped_unit_for_{q.question_type.lower()}")
                unit = None

        answers_norm[qid] = {
            "answer": ans2,
            "unit": unit,
            "evidence": evidence,
            "confidence": conf,
        }

    notes_raw = parsed.get("notes")
    notes = str(notes_raw) if notes_raw is not None else ""
    return {"answers_by_id": answers_norm, "notes": notes}, issues


def _append_jsonl(path: Path, obj: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Normalize OpenAI batch questionnaire results into a strict schema.")
    ap.add_argument("--request-dir", required=True, help="Questionnaire request dir containing mapping_shard*.jsonl")
    ap.add_argument("--results-dir", required=True, help="Results dir containing openai_results_shard*.jsonl")
    ap.add_argument("--questions-xlsx", required=True, help="Path to Questions.xlsx")
    ap.add_argument("--questions-processed-sheet", default="Processed Info", help="Processed sheet name")
    ap.add_argument("--questions-input-info-sheet", default="Input Info", help="Input Info sheet name")
    ap.add_argument(
        "--out-dir",
        default=None,
        help="Directory to write normalized.jsonl + invalid_rows.jsonl (default: <results-dir>)",
    )
    ap.add_argument("--max-lines", type=int, default=0, help="If >0, only process first N lines (debug)")
    return ap.parse_args()


def main() -> None:
    args = parse_args()
    request_dir = Path(expanduser(args.request_dir)).resolve()
    results_dir = Path(expanduser(args.results_dir)).resolve()
    out_dir = Path(expanduser(args.out_dir)).resolve() if args.out_dir else results_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    xlsx = Path(expanduser(args.questions_xlsx)).resolve()
    questions = _load_questions_xlsx(
        xlsx_path=xlsx,
        processed_sheet=str(args.questions_processed_sheet),
        input_info_sheet=str(args.questions_input_info_sheet),
    )
    expected_ids = [q.id for q in questions]

    mapping_by_id = _read_mapping(request_dir)

    result_paths = sorted(results_dir.glob("openai_results_shard*.jsonl"))
    if not result_paths:
        raise SystemExit(f"No openai_results_shard*.jsonl found in {results_dir}")

    out_norm = out_dir / "normalized.jsonl"
    out_bad = out_dir / "invalid_rows.jsonl"
    # Overwrite for determinism.
    out_norm.write_text("", encoding="utf-8")
    out_bad.write_text("", encoding="utf-8")

    status_code_counts: Counter[str] = Counter()
    body_status_counts: Counter[str] = Counter()
    parse_fail = 0
    normalized_ok = 0
    invalid = 0
    issue_counts: Counter[str] = Counter()
    non_null_counts: list[int] = []

    total_lines = 0
    max_lines = int(args.max_lines) if args.max_lines else 0

    for rp in result_paths:
        shard_s = rp.name[len("openai_results_shard") : -len(".jsonl")]
        shard = int(shard_s) if shard_s.isdigit() else -1
        for raw in rp.read_text(encoding="utf-8").splitlines():
            if max_lines and total_lines >= max_lines:
                break
            total_lines += 1
            line = raw.strip()
            if not line:
                continue
            obj = json.loads(line)
            custom_id = obj.get("custom_id")
            if not isinstance(custom_id, str) or not custom_id:
                continue

            resp = obj.get("response") or {}
            if not isinstance(resp, dict):
                status_code_counts["<missing_response>"] += 1
                invalid += 1
                _append_jsonl(out_bad, {"custom_id": custom_id, "shard": shard, "reason": "missing_response"})
                continue

            sc = resp.get("status_code")
            status_code_counts[str(sc)] += 1
            if sc != 200:
                invalid += 1
                _append_jsonl(
                    out_bad,
                    {"custom_id": custom_id, "shard": shard, "reason": f"status_code_{sc}", "response": resp},
                )
                continue

            body = resp.get("body")
            if not isinstance(body, dict):
                parse_fail += 1
                invalid += 1
                _append_jsonl(out_bad, {"custom_id": custom_id, "shard": shard, "reason": "missing_body"})
                continue

            body_status = str(body.get("status"))
            body_status_counts[body_status] += 1
            if body_status != "completed":
                invalid += 1
                _append_jsonl(out_bad, {"custom_id": custom_id, "shard": shard, "reason": f"body_status_{body_status}"})
                continue

            output_text = _extract_openai_output_text(body)
            if not output_text.strip():
                parse_fail += 1
                invalid += 1
                _append_jsonl(out_bad, {"custom_id": custom_id, "shard": shard, "reason": "empty_output_text"})
                continue

            try:
                parsed = _coerce_json(output_text)
            except Exception as exc:  # noqa: BLE001
                parse_fail += 1
                invalid += 1
                _append_jsonl(
                    out_bad,
                    {
                        "custom_id": custom_id,
                        "shard": shard,
                        "reason": "json_parse_error",
                        "error": str(exc),
                    },
                )
                continue

            normalized, issues = _normalize_one(custom_id=custom_id, parsed=parsed, questions=questions)
            if normalized is None:
                invalid += 1
                _append_jsonl(out_bad, {"custom_id": custom_id, "shard": shard, "reason": "cannot_normalize"})
                continue

            for iss in issues:
                issue_counts[iss] += 1

            # Non-null density (useful for downstream expectations).
            answers_by_id = normalized["answers_by_id"]
            nn = 0
            for qid in expected_ids:
                v = answers_by_id.get(qid) or {}
                if isinstance(v, dict) and v.get("answer") is not None:
                    nn += 1
            non_null_counts.append(nn)

            mapping = mapping_by_id.get(custom_id) or {}
            out_row = {
                "custom_id": custom_id,
                "page_id": mapping.get("page_id"),
                "box_id": mapping.get("box_id"),
                "bbox": mapping.get("bbox"),
                "cls": mapping.get("cls"),
                "source_page_path": mapping.get("page_path"),
                "normalized": normalized,
                "issues": issues,
                "shard": shard,
            }
            _append_jsonl(out_norm, out_row)
            normalized_ok += 1

        if max_lines and total_lines >= max_lines:
            break

    _eprint(f"request_dir\t{request_dir}")
    _eprint(f"results_dir\t{results_dir}")
    _eprint(f"out_norm\t{out_norm}")
    _eprint(f"out_bad\t{out_bad}")
    _eprint(f"question_count\t{len(questions)}")
    _eprint(f"total_lines\t{total_lines}")
    _eprint(f"status_code_counts\t{json.dumps(dict(status_code_counts), ensure_ascii=False)}")
    _eprint(f"body_status_counts\t{json.dumps(dict(body_status_counts), ensure_ascii=False)}")
    _eprint(f"parse_fail\t{parse_fail}")
    _eprint(f"normalized_ok\t{normalized_ok}")
    _eprint(f"invalid\t{invalid}")

    if non_null_counts:
        _eprint(f"non_null_any\t{sum(1 for n in non_null_counts if n > 0)}")
        _eprint(f"non_null_mean\t{statistics.mean(non_null_counts):.3f}")
        _eprint(f"non_null_p50\t{statistics.median(non_null_counts):.3f}")
        _eprint(f"non_null_max\t{max(non_null_counts)}")

    if issue_counts:
        top = dict(issue_counts.most_common(20))
        _eprint(f"top_issues\t{json.dumps(top, ensure_ascii=False)}")


if __name__ == "__main__":
    main()

