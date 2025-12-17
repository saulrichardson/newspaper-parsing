#!/usr/bin/env python3
"""
Run the ordinance/amendment questionnaire step synchronously (no Batch API).

Inputs:
  - Zoning box-classifier artifacts (OpenAI batch download):
      * mapping_shard*.jsonl (from the classifier request dir)
      * openai_results_shard*.jsonl (from the classifier results dir)
  - Per-page *.vlm.json OCR outputs referenced by mapping rows (page_path + box_id)
  - Questions.xlsx (Processed Info + Input Info)

Selection:
  - By default, only boxes classified as ordinance/amendment:
      full_ordinance, amendment_substantial, amendment_targeted

Output:
  - Per-box JSON outputs under:
      <output_dir>/pages/<page_id>/box_<box_id>.questionnaire.json
  - Append-only manifest:
      <output_dir>/manifest.jsonl

Key points:
  - Uses OpenAI synchronous Responses API (NOT batch).
  - Does NOT use OpenAI-side JSON-schema enforcement; instead we:
      1) instruct the model via prompt
      2) validate the returned JSON locally
      3) retry with explicit correction instructions on validation failure
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
import time
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from glob import glob
from os.path import expanduser
from pathlib import Path
from typing import Any, Literal


ZoningLabel = Literal[
    "full_ordinance",
    "amendment_substantial",
    "amendment_targeted",
    "public_hearing",
    "unrelated",
]


def _eprint(msg: str) -> None:
    print(msg, file=sys.stderr)


def _sha256(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _load_env_file(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        key = k.strip()
        val = v.strip().strip('"').strip("'")
        if key:
            env[key] = val
    return env


def _get_openai_key_candidates() -> list[tuple[str, str]]:
    # Prefer repo root .env if present (consistent with other scripts in this repo).
    file_env: dict[str, str] = {}
    for p in [Path(".env"), Path(__file__).resolve().parents[1] / ".env"]:
        if p.is_file():
            file_env = _load_env_file(p)
            break

    out: list[tuple[str, str]] = []
    for name in ["PROJECT_OPENAI_KEY", "OPENAI_API_KEY", "OPENAI_KEY"]:
        val = os.environ.get(name) or file_env.get(name)
        if not val:
            continue
        out.append((name, val))

    seen: set[str] = set()
    uniq: list[tuple[str, str]] = []
    for src, key in out:
        if key in seen:
            continue
        seen.add(key)
        uniq.append((src, key))

    if not uniq:
        raise SystemExit(
            "Missing OpenAI key. Set PROJECT_OPENAI_KEY (preferred) or OPENAI_API_KEY/OPENAI_KEY in env, "
            "or add them to .env in repo root."
        )
    return uniq


def _is_insufficient_quota_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    return ("insufficient_quota" in msg) or ("exceeded your current quota" in msg)


def _is_auth_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    return ("invalid_api_key" in msg) or ("401" in msg) or ("403" in msg) or ("unauthorized" in msg)


def _coerce_json(text: str) -> dict[str, Any]:
    """Accept raw JSON or JSON wrapped in ```json fences, scrub control chars."""
    if text is None:
        raise ValueError("empty response text")
    stripped = str(text).strip()
    if stripped.startswith("```"):
        lines = stripped.splitlines()
        lines = lines[1:]
        if lines and lines[-1].strip().startswith("```"):
            lines = lines[:-1]
        stripped = "\n".join(lines).strip()
    try:
        return json.loads(stripped)
    except json.JSONDecodeError as exc:
        import re

        cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", stripped)
        try:
            return json.loads(cleaned)
        except json.JSONDecodeError:
            if "Invalid control character" in str(exc):
                escaped = cleaned.replace("\n", "\\n")
                return json.loads(escaped)
        raise


def _extract_openai_output_text(body: dict[str, Any]) -> str:
    out = body.get("output")
    if not isinstance(out, list):
        return ""
    parts: list[str] = []
    for item in out:
        if not isinstance(item, dict):
            continue
        content = item.get("content") or []
        if not isinstance(content, list):
            continue
        for c in content:
            if isinstance(c, dict) and c.get("type") == "output_text":
                parts.append(str(c.get("text") or ""))
    return "".join(parts)


def _parse_shard_selector(selector: str) -> set[int]:
    selector = selector.strip()
    if not selector:
        return set()
    if "-" in selector and "," not in selector:
        a, b = selector.split("-", 1)
        start = int(a)
        end = int(b)
        if end < start:
            raise SystemExit(f"Invalid shard range: {selector}")
        return set(range(start, end + 1))
    parts = [p.strip() for p in selector.split(",") if p.strip()]
    return {int(p) for p in parts}


@dataclass(frozen=True)
class Question:
    id: str
    question_type: str
    full_question: str
    possible_answers: list[str] | None


def _load_questions_xlsx(*, xlsx_path: Path, processed_sheet: str, input_info_sheet: str) -> list[Question]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise SystemExit(
            "Missing dependency for reading .xlsx. Install openpyxl, e.g. `pip install openpyxl`."
        ) from exc

    if not xlsx_path.is_file():
        raise SystemExit(f"Questions workbook not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if processed_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found in workbook: {processed_sheet!r} (have {wb.sheetnames})")
    if input_info_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found in workbook: {input_info_sheet!r} (have {wb.sheetnames})")

    ws_proc = wb[processed_sheet]
    ws_in = wb[input_info_sheet]

    def _hdr_map(ws) -> dict[str, int]:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        out: dict[str, int] = {}
        for i, h in enumerate(headers, start=1):
            if h is None:
                continue
            key = str(h).strip()
            if not key:
                continue
            out[key] = i
        return out

    proc_idx = _hdr_map(ws_proc)
    in_idx = _hdr_map(ws_in)

    required_proc = {"ID", "Full Question", "Include", "Question Type"}
    missing_proc = sorted(required_proc - set(proc_idx))
    if missing_proc:
        raise SystemExit(f"Processed sheet missing required headers: {missing_proc}")

    # Build possible-answers lookup keyed by stringified ID.
    possible_by_id: dict[str, list[str]] = {}
    if "ID" in in_idx and "Possible Answers" in in_idx:
        for r in range(2, ws_in.max_row + 1):
            qid_raw = ws_in.cell(r, in_idx["ID"]).value
            if qid_raw is None:
                continue
            qid = str(qid_raw).strip()
            if not qid:
                continue
            poss_raw = ws_in.cell(r, in_idx["Possible Answers"]).value
            if poss_raw is None:
                continue
            poss_s = str(poss_raw).strip()
            if not poss_s:
                continue
            options = [p.strip() for p in poss_s.split(";") if p.strip()]
            if options:
                possible_by_id[qid] = options

    questions: list[Question] = []
    for r in range(2, ws_proc.max_row + 1):
        include_raw = ws_proc.cell(r, proc_idx["Include"]).value
        if str(include_raw).strip().lower() != "yes":
            continue

        qid_raw = ws_proc.cell(r, proc_idx["ID"]).value
        if qid_raw is None:
            continue
        qid = str(qid_raw).strip()
        if not qid:
            continue

        qtype_raw = ws_proc.cell(r, proc_idx["Question Type"]).value
        qtype = str(qtype_raw).strip() if qtype_raw is not None else ""
        if qtype not in {"Binary", "Categorical", "Numerical", "Continuous"}:
            raise SystemExit(f"Unsupported Question Type {qtype!r} for ID={qid} (row {r})")

        full_raw = ws_proc.cell(r, proc_idx["Full Question"]).value
        full_q = str(full_raw).strip() if full_raw is not None else ""
        if not full_q:
            raise SystemExit(f"Missing Full Question for ID={qid} (row {r})")

        questions.append(
            Question(
                id=qid,
                question_type=qtype,
                full_question=full_q,
                possible_answers=possible_by_id.get(qid),
            )
        )

    if not questions:
        raise SystemExit(
            f"No included questions found in sheet {processed_sheet!r}. "
            "Expected at least one row where Include == 'Yes'."
        )
    return questions


def _questions_payload_for_prompt(questions: list[Question]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for q in questions:
        obj: dict[str, Any] = {
            "id": q.id,
            "question_type": q.question_type,
            "question": q.full_question,
        }
        if q.possible_answers:
            obj["options"] = q.possible_answers
        out.append(obj)
    return out


def _validate_output(*, parsed: dict[str, Any], questions: list[Question]) -> list[str]:
    errs: list[str] = []
    if not isinstance(parsed, dict):
        return ["output is not a JSON object"]

    answers_by_id = parsed.get("answers_by_id")
    if not isinstance(answers_by_id, dict):
        errs.append("missing/invalid answers_by_id (must be an object)")
        return errs

    expected_ids = [q.id for q in questions]
    expected_set = set(expected_ids)
    got_set = set(str(k) for k in answers_by_id.keys())
    missing = sorted(expected_set - got_set)
    extra = sorted(got_set - expected_set)
    if missing:
        errs.append(f"missing question ids: {missing[:20]}{' (+more)' if len(missing) > 20 else ''}")
    if extra:
        errs.append(f"extra question ids: {extra[:20]}{' (+more)' if len(extra) > 20 else ''}")

    q_by_id = {q.id: q for q in questions}

    def _as_float(v) -> float | None:
        if isinstance(v, (int, float)):
            return float(v)
        return None

    for qid in expected_ids:
        a = answers_by_id.get(qid)
        if not isinstance(a, dict):
            errs.append(f"{qid}: value must be an object")
            continue

        # Required keys.
        for req_key in ["answer", "unit", "evidence", "confidence"]:
            if req_key not in a:
                errs.append(f"{qid}: missing key {req_key!r}")

        ans = a.get("answer")
        unit = a.get("unit")
        evidence = a.get("evidence")
        conf_raw = a.get("confidence")
        conf = _as_float(conf_raw)

        if conf is None:
            errs.append(f"{qid}: confidence must be a number 0..1")
        elif not (0.0 <= conf <= 1.0):
            errs.append(f"{qid}: confidence out of range 0..1: {conf}")

        # Normalize unit/evidence empties for validation convenience.
        if unit == "":
            unit = None
        if evidence == "":
            evidence = None

        if ans is None:
            if unit is not None:
                errs.append(f"{qid}: unit must be null when answer is null")
            if evidence is not None:
                errs.append(f"{qid}: evidence must be null when answer is null")
            if conf is not None and conf != 0.0:
                errs.append(f"{qid}: confidence must be 0 when answer is null")
            continue

        # answer non-null: evidence required
        if not isinstance(evidence, str) or not evidence.strip():
            errs.append(f"{qid}: evidence required (non-empty string) when answer is non-null")
        elif len(evidence) > 240:
            errs.append(f"{qid}: evidence too long (>240 chars)")

        q = q_by_id[qid]
        qtype = q.question_type
        if qtype == "Binary":
            if not isinstance(ans, bool):
                errs.append(f"{qid}: Binary answer must be true/false/null")
            if unit is not None:
                errs.append(f"{qid}: unit must be null for Binary questions")
        elif qtype in {"Numerical", "Continuous"}:
            if not isinstance(ans, (int, float)):
                errs.append(f"{qid}: {qtype} answer must be a number or null")
            if unit is not None and not isinstance(unit, str):
                errs.append(f"{qid}: unit must be string or null for numeric questions")
        elif qtype == "Categorical":
            if not isinstance(ans, str):
                errs.append(f"{qid}: Categorical answer must be a string or null")
            if unit is not None:
                errs.append(f"{qid}: unit must be null for Categorical questions")
            if q.possible_answers and isinstance(ans, str) and ans not in q.possible_answers:
                errs.append(f"{qid}: categorical answer not in options: {ans!r}")
        else:
            errs.append(f"{qid}: unknown question_type {qtype!r}")

    return errs


def _collect_mapping_paths(request_dir: Path, *, want: set[int] | None) -> list[tuple[int, Path]]:
    paths: list[tuple[int, Path]] = []
    for p in sorted(Path(x) for x in glob(str(request_dir / "mapping_shard*.jsonl"))):
        name = p.name
        if not name.startswith("mapping_shard") or not name.endswith(".jsonl"):
            continue
        shard_s = name[len("mapping_shard") : -len(".jsonl")]
        try:
            shard = int(shard_s)
        except ValueError:
            continue
        if want is not None and shard not in want:
            continue
        paths.append((shard, p))
    return paths


def _collect_openai_result_paths(results_dir: Path, *, want: set[int] | None) -> list[tuple[int, Path]]:
    paths: list[tuple[int, Path]] = []
    for p in sorted(Path(x) for x in glob(str(results_dir / "openai_results_shard*.jsonl"))):
        name = p.name
        if not name.startswith("openai_results_shard") or not name.endswith(".jsonl"):
            continue
        shard_s = name[len("openai_results_shard") : -len(".jsonl")]
        try:
            shard = int(shard_s)
        except ValueError:
            continue
        if want is not None and shard not in want:
            continue
        paths.append((shard, p))
    return paths


def _load_box_transcript(*, page_path: Path, box_id: int) -> tuple[str, dict[str, Any]]:
    data = json.loads(page_path.read_text(encoding="utf-8"))
    boxes = data.get("boxes")
    if not isinstance(boxes, list):
        raise ValueError(f"Invalid page JSON (missing boxes list): {page_path}")
    for b in boxes:
        if not isinstance(b, dict):
            continue
        if b.get("id") != box_id:
            continue
        status = b.get("status")
        transcript = (b.get("transcript") or "").strip()
        if status != "ok":
            raise ValueError(f"Box {box_id} status is {status!r}, expected 'ok' ({page_path})")
        if not transcript:
            raise ValueError(f"Box {box_id} has empty transcript despite ok status ({page_path})")
        stats = {
            "source_model": data.get("model"),
            "box_cls": b.get("class") or b.get("cls"),
            "box_text_chars": len(transcript),
            "box_text_sha256": _sha256(transcript),
        }
        return transcript, stats
    raise ValueError(f"Box id {box_id} not found in {page_path}")


def _append_jsonl(path: Path, obj: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def _atomic_write_json(path: Path, obj: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(obj, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    tmp.replace(path)


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Run ordinance questionnaire synchronously (OpenAI).")
    ap.add_argument("--classification-request-dir", required=True, help="Classifier request dir with mapping_shard*.jsonl")
    ap.add_argument(
        "--classification-results-dir",
        required=True,
        help="Classifier results dir with openai_results_shard*.jsonl",
    )
    ap.add_argument("--output-dir", required=True, help="Directory to write per-box questionnaire outputs")

    ap.add_argument("--questions-xlsx", required=True, help="Path to Questions.xlsx")
    ap.add_argument("--questions-processed-sheet", default="Processed Info", help="Workbook sheet for processed questions")
    ap.add_argument("--questions-input-info-sheet", default="Input Info", help="Workbook sheet for possible answers")
    ap.add_argument("--prompt-path", default="prompts/ordinance_box_questionnaire_prompt_text.txt", help="Prompt file path")

    ap.add_argument(
        "--include-labels",
        default="full_ordinance,amendment_substantial,amendment_targeted",
        help="Comma-separated zoning labels to include from classifier outputs",
    )
    ap.add_argument("--use-present-flags", action="store_true", help="Also include boxes based on present.* flags")
    ap.add_argument("--min-confidence", type=float, default=0.0, help="Minimum classifier confidence to include")

    ap.add_argument("--shards", default=None, help="Optional classifier shard selector, e.g. '0-10' or '0,1,2'")
    ap.add_argument(
        "--allow-partial-results",
        action="store_true",
        help="Allow running even if some mapping_shardNNN have no corresponding openai_results_shardNNN.",
    )
    ap.add_argument("--max-boxes", type=int, default=None, help="Stop after processing this many selected boxes")
    ap.add_argument("--skip-existing", action="store_true", help="Skip boxes with existing output file")

    ap.add_argument("--openai-model", default="gpt-5-nano", help="OpenAI model name")
    ap.add_argument(
        "--openai-reasoning-effort",
        choices=["none", "minimal", "low", "medium", "high", "xhigh"],
        default="medium",
        help="OpenAI reasoning.effort (set 'none' to omit)",
    )
    ap.add_argument("--timeout", type=float, default=180.0, help="Timeout seconds per OpenAI call")
    ap.add_argument("--max-retries", type=int, default=3, help="Retries on parse/validation failure per box")
    ap.add_argument("--retry-sleep", type=float, default=0.5, help="Seconds to sleep between retries (backoff multiplier applied)")
    ap.add_argument("--dry-run", action="store_true", help="Scan + count selected boxes; do not call OpenAI")

    return ap.parse_args()


def main() -> None:
    args = parse_args()

    cls_request_dir = Path(expanduser(args.classification_request_dir)).resolve()
    cls_results_dir = Path(expanduser(args.classification_results_dir)).resolve()
    out_dir = Path(expanduser(args.output_dir)).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = out_dir / "manifest.jsonl"

    questions_xlsx = Path(expanduser(args.questions_xlsx)).resolve()
    prompt_path = Path(expanduser(args.prompt_path)).resolve()
    if not prompt_path.is_file():
        raise SystemExit(f"Prompt file not found: {prompt_path}")
    prompt_text = prompt_path.read_text(encoding="utf-8").strip()

    include_labels = {s.strip() for s in str(args.include_labels).split(",") if s.strip()}
    if not include_labels:
        raise SystemExit("--include-labels produced empty set")

    want_shards: set[int] | None = None
    if args.shards:
        want_shards = _parse_shard_selector(args.shards)
        if not want_shards:
            raise SystemExit(f"--shards parsed to empty set: {args.shards!r}")

    mapping_paths = _collect_mapping_paths(cls_request_dir, want=want_shards)
    if not mapping_paths:
        raise SystemExit(f"No mapping_shard*.jsonl found in {cls_request_dir}")

    result_paths = _collect_openai_result_paths(cls_results_dir, want=want_shards)
    if not result_paths:
        raise SystemExit(f"No openai_results_shard*.jsonl found in {cls_results_dir}")

    mapping_shards = {s for s, _ in mapping_paths}
    result_shards = {s for s, _ in result_paths}
    missing = sorted(mapping_shards - result_shards)
    if missing and not bool(args.allow_partial_results):
        preview = ",".join(str(s) for s in missing[:20])
        more = "" if len(missing) <= 20 else f" (+{len(missing) - 20} more)"
        raise SystemExit(
            "Classifier results are incomplete; refusing to run questionnaire.\n"
            f"Missing openai_results_shardNNN.jsonl for shards: {preview}{more}\n"
            "Run the downloader until all shards are present, OR re-run with either:\n"
            "  - --shards <subset> (run only completed shards), or\n"
            "  - --allow-partial-results (explicitly accept partial run)."
        )

    questions = _load_questions_xlsx(
        xlsx_path=questions_xlsx,
        processed_sheet=str(args.questions_processed_sheet),
        input_info_sheet=str(args.questions_input_info_sheet),
    )
    questions_payload = _questions_payload_for_prompt(questions)
    questions_json = json.dumps(questions_payload, ensure_ascii=False, separators=(",", ":"))
    question_set_sha = _sha256(json.dumps(questions_payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")))

    # Load mapping in memory: custom_id -> mapping row
    mapping_by_id: dict[str, dict[str, Any]] = {}
    for shard, mp in mapping_paths:
        with mp.open("r", encoding="utf-8") as f:
            for raw in f:
                line = raw.strip()
                if not line:
                    continue
                obj = json.loads(line)
                key = obj.get("id") or obj.get("custom_id")
                if not isinstance(key, str) or not key:
                    continue
                mapping_by_id[key] = obj

    # OpenAI client(s) (project key first, then fallbacks).
    from openai import OpenAI  # imported late

    clients: list[tuple[str, Any]] = []
    for src, key in _get_openai_key_candidates():
        clients.append((src, OpenAI(api_key=key, timeout=args.timeout)))

    min_conf = float(args.min_confidence)
    label_counts = Counter()

    selected = 0
    processed = 0
    ok = 0
    err = 0
    skipped_existing = 0
    skipped_no_mapping = 0
    skipped_http_error = 0
    skipped_not_completed = 0
    skipped_parse_error = 0

    def build_base_prompt(*, transcript: str) -> str:
        # Deterministic and easy to validate.
        return (
            f"{prompt_text}\n\n"
            f"QUESTIONS_JSON:\n{questions_json}\n\n"
            "OCR_TEXT:\n"
            f"{transcript}\n"
        )

    # Cache last loaded page to avoid re-reading JSON for adjacent boxes on the same page.
    last_page_path: Path | None = None
    last_page_boxes: dict[int, tuple[str, dict[str, Any]]] = {}

    def get_transcript_for(mapping_row: dict[str, Any]) -> tuple[str, dict[str, Any]]:
        nonlocal last_page_path, last_page_boxes
        page_path_raw = mapping_row.get("page_path")
        box_id = mapping_row.get("box_id")
        if not isinstance(page_path_raw, str) or not page_path_raw:
            raise ValueError("mapping missing page_path")
        if not isinstance(box_id, int):
            try:
                box_id = int(box_id)
            except Exception as exc:
                raise ValueError(f"mapping missing/invalid box_id: {box_id!r}") from exc

        page_path = Path(expanduser(page_path_raw)).resolve()
        if not page_path.is_file():
            raise ValueError(f"page_path not found: {page_path}")

        if last_page_path != page_path:
            data = json.loads(page_path.read_text(encoding="utf-8"))
            boxes = data.get("boxes")
            if not isinstance(boxes, list):
                raise ValueError(f"Invalid page JSON (missing boxes list): {page_path}")
            last_page_boxes = {}
            for b in boxes:
                if not isinstance(b, dict):
                    continue
                bid = b.get("id")
                if not isinstance(bid, int):
                    continue
                status = b.get("status")
                transcript = (b.get("transcript") or "").strip()
                if status == "ok" and transcript:
                    stats = {
                        "source_model": data.get("model"),
                        "box_cls": b.get("class") or b.get("cls"),
                        "box_text_chars": len(transcript),
                        "box_text_sha256": _sha256(transcript),
                    }
                    last_page_boxes[bid] = (transcript, stats)
            last_page_path = page_path

        if int(box_id) not in last_page_boxes:
            return _load_box_transcript(page_path=page_path, box_id=int(box_id))
        return last_page_boxes[int(box_id)]

    def out_path_for(*, page_id: str, box_id: int) -> Path:
        return out_dir / "pages" / page_id / f"box_{int(box_id)}.questionnaire.json"

    def should_include(*, label: Any, conf_f: float, present: Any) -> bool:
        if isinstance(label, str) and label in include_labels and conf_f >= min_conf:
            return True
        if not bool(args.use_present_flags):
            return False
        if not isinstance(present, dict) or conf_f < min_conf:
            return False
        for k in ["full_ordinance", "amendment_substantial", "amendment_targeted"]:
            if present.get(k) is True and k in include_labels:
                return True
        return False

    def call_openai_with_retry(*, prompt: str, expected_ids: list[str]) -> tuple[dict[str, Any], dict[str, Any]]:
        # Returns (validated_output, meta)
        # meta includes response_id, auth_source, attempts, raw_output_text (last)
        attempts = 0
        last_errs: list[str] | None = None
        last_raw: str | None = None
        last_resp_id: str | None = None
        last_auth: str | None = None

        while attempts < int(args.max_retries):
            attempts += 1
            repair_msg = ""
            if last_errs:
                # Keep this short-ish; the base prompt already contains schema + QUESTIONS_JSON.
                err_block = "\n".join(f"- {e}" for e in last_errs[:30])
                repair_msg = (
                    "\n\nYour previous output was invalid.\n"
                    "Fix the JSON to satisfy the exact schema and constraints.\n"
                    f"Validation errors:\n{err_block}\n"
                    "Output JSON only.\n"
                )

            # Try keys in preference order (project key first).
            resp = None
            auth_src = None
            last_exc: Exception | None = None
            for src, client in clients:
                try:
                    body: dict[str, Any] = {
                        "model": args.openai_model,
                        "input": [
                            {
                                "role": "user",
                                "content": [{"type": "input_text", "text": prompt + repair_msg}],
                            }
                        ],
                    }
                    if args.openai_reasoning_effort and args.openai_reasoning_effort != "none":
                        body["reasoning"] = {"effort": args.openai_reasoning_effort}
                    resp = client.responses.create(**body)
                    auth_src = src
                    break
                except Exception as exc:  # noqa: BLE001
                    last_exc = exc
                    if _is_insufficient_quota_error(exc) or _is_auth_error(exc):
                        continue
                    raise

            if resp is None:
                raise RuntimeError(str(last_exc) if last_exc else "OpenAI request failed for all keys")

            last_resp_id = getattr(resp, "id", None)
            last_auth = auth_src
            raw_text = getattr(resp, "output_text", None)
            if not raw_text:
                last_raw = None
                last_errs = ["missing output_text in OpenAI response"]
            else:
                last_raw = str(raw_text)
                try:
                    parsed = _coerce_json(last_raw)
                    errs = _validate_output(parsed=parsed, questions=questions)
                    # Extra strict: ensure we got exactly the expected ids set (prompt says so).
                    if not errs:
                        return parsed, {
                            "response_id": last_resp_id,
                            "auth_source": last_auth,
                            "attempts": attempts,
                            "raw_output_text": last_raw,
                        }
                    last_errs = errs
                except Exception as exc:  # noqa: BLE001
                    last_errs = [f"json_parse_error: {exc}"]

            sleep_s = float(args.retry_sleep) * attempts
            time.sleep(sleep_s)

        # Exhausted retries.
        raise RuntimeError(
            "Failed to get valid questionnaire JSON after retries. "
            f"last_response_id={last_resp_id} auth_source={last_auth} errors={last_errs}"
        )

    expected_ids = [q.id for q in questions]

    # Iterate classifier results and process selected boxes.
    for shard, rp in result_paths:
        with rp.open("r", encoding="utf-8") as f:
            for raw in f:
                if args.max_boxes is not None and processed >= int(args.max_boxes):
                    break

                line = raw.strip()
                if not line:
                    continue
                obj = json.loads(line)
                custom_id = obj.get("custom_id")
                if not isinstance(custom_id, str) or not custom_id:
                    continue

                mapping_row = mapping_by_id.get(custom_id)
                if mapping_row is None:
                    skipped_no_mapping += 1
                    continue

                resp = obj.get("response") or {}
                if not isinstance(resp, dict):
                    skipped_http_error += 1
                    continue
                if resp.get("status_code") != 200:
                    skipped_http_error += 1
                    continue
                body = resp.get("body") if isinstance(resp.get("body"), dict) else None
                if not isinstance(body, dict):
                    skipped_parse_error += 1
                    continue
                if body.get("status") != "completed":
                    skipped_not_completed += 1
                    continue

                output_text = _extract_openai_output_text(body)
                if not output_text.strip():
                    skipped_parse_error += 1
                    continue

                try:
                    cls_parsed = _coerce_json(output_text)
                except Exception:
                    skipped_parse_error += 1
                    continue

                label = cls_parsed.get("label")
                conf = cls_parsed.get("confidence")
                present = cls_parsed.get("present")

                if isinstance(label, str):
                    label_counts[label] += 1

                try:
                    conf_f = float(conf) if isinstance(conf, (int, float)) else 0.0
                except Exception:
                    conf_f = 0.0

                if not should_include(label=label, conf_f=conf_f, present=present):
                    continue

                selected += 1

                page_id = mapping_row.get("page_id")
                box_id = mapping_row.get("box_id")
                if not isinstance(page_id, str) or not page_id:
                    _eprint(f"warn: mapping missing page_id for {custom_id}; skipping")
                    continue
                if not isinstance(box_id, int):
                    try:
                        box_id = int(box_id)
                    except Exception:
                        _eprint(f"warn: mapping missing/invalid box_id for {custom_id}; skipping")
                        continue

                out_path = out_path_for(page_id=page_id, box_id=int(box_id))
                if args.skip_existing and out_path.exists():
                    skipped_existing += 1
                    continue

                # Load transcript
                try:
                    transcript, tstats = get_transcript_for(mapping_row)
                except Exception as exc:  # noqa: BLE001
                    err += 1
                    rec = {
                        "custom_id": custom_id,
                        "page_id": page_id,
                        "box_id": int(box_id),
                        "status": "error",
                        "error": {"message": str(exc)},
                        "stage": "load_transcript",
                        "mapping_row": {
                            "page_path": mapping_row.get("page_path"),
                            "bbox": mapping_row.get("bbox"),
                            "cls": mapping_row.get("cls"),
                        },
                        "classifier": {"label": label, "confidence": conf_f, "present": present},
                        "written_at": _now_iso(),
                    }
                    _append_jsonl(manifest_path, rec)
                    continue

                base_prompt = build_base_prompt(transcript=transcript)
                started_at = _now_iso()
                t0 = time.perf_counter()
                try:
                    if args.dry_run:
                        processed += 1
                        continue

                    parsed_out, meta = call_openai_with_retry(prompt=base_prompt, expected_ids=expected_ids)

                    finished_at = _now_iso()
                    duration_ms = (time.perf_counter() - t0) * 1000.0

                    # Write per-box output
                    out_obj = {
                        "custom_id": custom_id,
                        "page_id": page_id,
                        "box_id": int(box_id),
                        "page_path": mapping_row.get("page_path"),
                        "bbox": mapping_row.get("bbox"),
                        "cls": mapping_row.get("cls"),
                        "source_model": tstats.get("source_model"),
                        "questionnaire_model": args.openai_model,
                        "openai_auth_source": meta.get("auth_source"),
                        "openai_response_id": meta.get("response_id"),
                        "prompt_path": str(prompt_path),
                        "questions_xlsx": str(questions_xlsx),
                        "questions_processed_sheet": str(args.questions_processed_sheet),
                        "questions_input_info_sheet": str(args.questions_input_info_sheet),
                        "question_count": len(questions),
                        "question_set_sha256": question_set_sha,
                        "classifier": {"label": label, "confidence": conf_f, "present": present},
                        "input_stats": {
                            "box_text_chars": tstats.get("box_text_chars"),
                            "box_text_sha256": tstats.get("box_text_sha256"),
                        },
                        "started_at": started_at,
                        "finished_at": finished_at,
                        "attempts": int(meta.get("attempts") or 0),
                        "duration_ms": duration_ms,
                        "output": parsed_out,
                    }
                    _atomic_write_json(out_path, out_obj)

                    ok += 1
                    processed += 1
                    _append_jsonl(
                        manifest_path,
                        {
                            "custom_id": custom_id,
                            "page_id": page_id,
                            "box_id": int(box_id),
                            "status": "ok",
                            "output_path": str(out_path),
                            "questionnaire_model": args.openai_model,
                            "openai_auth_source": meta.get("auth_source"),
                            "openai_response_id": meta.get("response_id"),
                            "attempts": int(meta.get("attempts") or 0),
                            "duration_ms": duration_ms,
                            "written_at": finished_at,
                        },
                    )
                except Exception as exc:  # noqa: BLE001
                    finished_at = _now_iso()
                    duration_ms = (time.perf_counter() - t0) * 1000.0
                    err += 1
                    _append_jsonl(
                        manifest_path,
                        {
                            "custom_id": custom_id,
                            "page_id": page_id,
                            "box_id": int(box_id),
                            "status": "error",
                            "error": {"message": str(exc)},
                            "questionnaire_model": args.openai_model,
                            "started_at": started_at,
                            "finished_at": finished_at,
                            "duration_ms": duration_ms,
                        },
                    )

        if args.max_boxes is not None and processed >= int(args.max_boxes):
            break

    print(f"output_dir\t{out_dir}")
    print(f"questions_xlsx\t{questions_xlsx}")
    print(f"question_count\t{len(questions)}")
    print(f"question_set_sha256\t{question_set_sha}")
    print(f"classifier_results_shards_seen\t{len(result_paths)}")
    print(f"classifier_mapping_keys_loaded\t{len(mapping_by_id)}")
    print(f"include_labels\t{sorted(include_labels)}")
    print(f"min_confidence\t{min_conf}")
    print(f"selected_boxes_seen\t{selected}")
    if args.dry_run:
        print("dry_run\ttrue")
    print(f"processed_boxes\t{processed}")
    print(f"ok\t{ok}")
    print(f"errors\t{err}")
    print(f"skipped_existing\t{skipped_existing}")
    print(f"skipped_no_mapping\t{skipped_no_mapping}")
    print(f"skipped_http_error\t{skipped_http_error}")
    print(f"skipped_not_completed\t{skipped_not_completed}")
    print(f"skipped_parse_error\t{skipped_parse_error}")
    if label_counts:
        print("classifier_label_counts\t" + json.dumps(dict(label_counts), ensure_ascii=False))
    print(f"manifest_path\t{manifest_path}")


if __name__ == "__main__":
    main()

