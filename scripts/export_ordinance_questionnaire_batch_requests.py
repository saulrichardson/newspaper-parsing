#!/usr/bin/env python3
"""
Export batch-request JSONLs for ordinance/amendment questionnaire answering at the BOX level.

Goal:
  - Start from a prior zoning box-classification run (OpenAI Batch results).
  - Select boxes whose zoning label indicates ordinance/amendment text:
      full_ordinance, amendment_substantial, amendment_targeted
    (optionally also include boxes where `present.*` flags indicate ordinance/amendment).
  - Load each selected box's OCR transcript from the original per-page *.vlm.json
    referenced by the classifier request mapping_shard*.jsonl.
  - Emit provider batch request JSONLs (OpenAI and/or Gemini) that answer a fixed
    questionnaire read from an Excel workbook.

Outputs (sharded JSONL in --output-dir):
  - mapping_shardNNN.jsonl
  - openai_requests_shardNNN.jsonl (optional)
  - gemini_requests_shardNNN.jsonl (optional)

This script does NOT submit anything; it only writes request shards.
Use scripts/submit_batch_shards.py (or the Slurm helpers) to submit.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import Counter
from dataclasses import dataclass
from glob import glob
from os.path import expanduser
from pathlib import Path
from typing import IO, Any, Iterable, Literal


Provider = Literal["openai", "gemini", "both"]
OpenAITextFormat = Literal["json_object", "json_schema", "none"]


def _eprint(msg: str) -> None:
    print(msg, file=sys.stderr)


def _sha256(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _coerce_json(text: str) -> dict[str, Any]:
    """Accept raw JSON or JSON wrapped in ```json fences, scrub control chars."""
    if text is None:
        raise ValueError("empty response text")
    stripped = str(text).strip()
    if stripped.startswith("```"):
        lines = stripped.splitlines()
        lines = lines[1:]
        if lines and lines[-1].strip().startswith("```"):
            lines = lines[:-1]
        stripped = "\n".join(lines).strip()
    try:
        return json.loads(stripped)
    except json.JSONDecodeError as exc:
        import re

        cleaned = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", stripped)
        try:
            return json.loads(cleaned)
        except json.JSONDecodeError:
            if "Invalid control character" in str(exc):
                escaped = cleaned.replace("\n", "\\n")
                return json.loads(escaped)
        raise


def _parse_shard_selector(selector: str) -> set[int]:
    selector = selector.strip()
    if not selector:
        return set()
    if "-" in selector and "," not in selector:
        a, b = selector.split("-", 1)
        start = int(a)
        end = int(b)
        if end < start:
            raise SystemExit(f"Invalid shard range: {selector}")
        return set(range(start, end + 1))
    parts = [p.strip() for p in selector.split(",") if p.strip()]
    return {int(p) for p in parts}


def _extract_openai_output_text(body: dict[str, Any]) -> str:
    # OpenAI Responses API returns `output: [{... content: [{type: output_text, text: ...}, ...]}]`.
    out = body.get("output")
    if not isinstance(out, list):
        return ""
    parts: list[str] = []
    for item in out:
        if not isinstance(item, dict):
            continue
        content = item.get("content") or []
        if not isinstance(content, list):
            continue
        for c in content:
            if isinstance(c, dict) and c.get("type") == "output_text":
                parts.append(str(c.get("text") or ""))
    return "".join(parts)


def _openai_text_config_json_object() -> dict[str, Any]:
    return {"format": {"type": "json_object"}}


def _questionnaire_openai_json_schema() -> dict[str, Any]:
    # NOTE: OpenAI strict JSON-schema mode requires additionalProperties:false for objects.
    return {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "answers": {
                "type": "array",
                "items": {
                    "type": "object",
                    "additionalProperties": False,
                    "properties": {
                        "id": {"type": "string"},
                        "question_type": {
                            "type": "string",
                            "enum": ["Binary", "Categorical", "Numerical", "Continuous"],
                        },
                        "answer": {"type": ["boolean", "string", "number", "null"]},
                        "unit": {"type": ["string", "null"]},
                        "evidence": {"type": ["string", "null"]},
                        "confidence": {"type": "number", "minimum": 0.0, "maximum": 1.0},
                    },
                    "required": ["id", "question_type", "answer", "unit", "evidence", "confidence"],
                },
            },
            "notes": {"type": "string"},
        },
        "required": ["answers", "notes"],
    }


def _openai_text_config(fmt: OpenAITextFormat) -> dict[str, Any] | None:
    if fmt == "none":
        return None
    if fmt == "json_object":
        return _openai_text_config_json_object()
    if fmt == "json_schema":
        return {
            "format": {
                "type": "json_schema",
                "name": "ordinance_questionnaire_output",
                "schema": _questionnaire_openai_json_schema(),
                "strict": True,
            }
        }
    raise ValueError(f"Unknown openai_text_format: {fmt}")


@dataclass(frozen=True)
class Question:
    id: str
    question_type: str
    full_question: str
    short_question: str | None
    possible_answers: list[str] | None


def _load_questions_xlsx(*, xlsx_path: Path, processed_sheet: str, input_info_sheet: str) -> list[Question]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise SystemExit(
            "Missing dependency for reading .xlsx. Install openpyxl, e.g. `pip install openpyxl`."
        ) from exc

    if not xlsx_path.is_file():
        raise SystemExit(f"Questions workbook not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if processed_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found in workbook: {processed_sheet!r} (have {wb.sheetnames})")
    if input_info_sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found in workbook: {input_info_sheet!r} (have {wb.sheetnames})")

    ws_proc = wb[processed_sheet]
    ws_in = wb[input_info_sheet]

    def _hdr_map(ws) -> dict[str, int]:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        out: dict[str, int] = {}
        for i, h in enumerate(headers, start=1):
            if h is None:
                continue
            key = str(h).strip()
            if not key:
                continue
            out[key] = i
        return out

    proc_idx = _hdr_map(ws_proc)
    in_idx = _hdr_map(ws_in)

    required_proc = {"ID", "Full Question", "Include", "Question Type"}
    missing_proc = sorted(required_proc - set(proc_idx))
    if missing_proc:
        raise SystemExit(f"Processed sheet missing required headers: {missing_proc}")

    # Optional in-sheet columns.
    proc_short_col = proc_idx.get("Short Question")

    # Build possible-answers lookup keyed by stringified ID.
    possible_by_id: dict[str, list[str]] = {}
    if "ID" in in_idx and "Possible Answers" in in_idx:
        for r in range(2, ws_in.max_row + 1):
            qid_raw = ws_in.cell(r, in_idx["ID"]).value
            if qid_raw is None:
                continue
            qid = str(qid_raw).strip()
            if not qid:
                continue
            poss_raw = ws_in.cell(r, in_idx["Possible Answers"]).value
            if poss_raw is None:
                continue
            poss_s = str(poss_raw).strip()
            if not poss_s:
                continue
            # Workbook uses ';' separated options.
            options = [p.strip() for p in poss_s.split(";") if p.strip()]
            if options:
                possible_by_id[qid] = options

    questions: list[Question] = []
    for r in range(2, ws_proc.max_row + 1):
        include_raw = ws_proc.cell(r, proc_idx["Include"]).value
        if str(include_raw).strip().lower() != "yes":
            continue

        qid_raw = ws_proc.cell(r, proc_idx["ID"]).value
        if qid_raw is None:
            continue
        qid = str(qid_raw).strip()
        if not qid:
            continue

        qtype_raw = ws_proc.cell(r, proc_idx["Question Type"]).value
        qtype = str(qtype_raw).strip() if qtype_raw is not None else ""
        if qtype not in {"Binary", "Categorical", "Numerical", "Continuous"}:
            raise SystemExit(f"Unsupported Question Type {qtype!r} for ID={qid} (row {r})")

        full_raw = ws_proc.cell(r, proc_idx["Full Question"]).value
        full_q = str(full_raw).strip() if full_raw is not None else ""
        if not full_q:
            raise SystemExit(f"Missing Full Question for ID={qid} (row {r})")

        short_q = None
        if proc_short_col is not None:
            short_raw = ws_proc.cell(r, proc_short_col).value
            if short_raw is not None and str(short_raw).strip():
                short_q = str(short_raw).strip()

        questions.append(
            Question(
                id=qid,
                question_type=qtype,
                full_question=full_q,
                short_question=short_q,
                possible_answers=possible_by_id.get(qid),
            )
        )

    if not questions:
        raise SystemExit(
            f"No included questions found in sheet {processed_sheet!r}. "
            "Expected at least one row where Include == 'Yes'."
        )
    return questions


def _questions_payload_for_prompt(questions: list[Question]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for q in questions:
        obj: dict[str, Any] = {
            "id": q.id,
            "question_type": q.question_type,
            "question": q.full_question,
        }
        if q.possible_answers:
            obj["options"] = q.possible_answers
        out.append(obj)
    return out


def _load_box_transcript(*, page_path: Path, box_id: int) -> tuple[str, dict[str, Any]]:
    """Return (transcript, stats). Raises on missing/unusable box."""
    data = json.loads(page_path.read_text(encoding="utf-8"))
    boxes = data.get("boxes")
    if not isinstance(boxes, list):
        raise ValueError(f"Invalid page JSON (missing boxes list): {page_path}")
    for b in boxes:
        if not isinstance(b, dict):
            continue
        if b.get("id") != box_id:
            continue
        status = b.get("status")
        transcript = (b.get("transcript") or "").strip()
        if status != "ok":
            raise ValueError(f"Box {box_id} status is {status!r}, expected 'ok' ({page_path})")
        if not transcript:
            raise ValueError(f"Box {box_id} has empty transcript despite ok status ({page_path})")
        stats = {
            "source_model": data.get("model"),
            "box_cls": b.get("class") or b.get("cls"),
            "box_text_chars": len(transcript),
            "box_text_sha256": _sha256(transcript),
        }
        return transcript, stats
    raise ValueError(f"Box id {box_id} not found in {page_path}")


def _collect_mapping_paths(request_dir: Path, *, want: set[int] | None) -> list[tuple[int, Path]]:
    # mapping_shardNNN.jsonl
    paths: list[tuple[int, Path]] = []
    for p in sorted(Path(x) for x in glob(str(request_dir / "mapping_shard*.jsonl"))):
        name = p.name
        if not name.startswith("mapping_shard") or not name.endswith(".jsonl"):
            continue
        shard_s = name[len("mapping_shard") : -len(".jsonl")]
        try:
            shard = int(shard_s)
        except ValueError:
            continue
        if want is not None and shard not in want:
            continue
        paths.append((shard, p))
    return paths


def _collect_openai_result_paths(results_dir: Path, *, want: set[int] | None) -> list[tuple[int, Path]]:
    # openai_results_shardNNN.jsonl
    paths: list[tuple[int, Path]] = []
    for p in sorted(Path(x) for x in glob(str(results_dir / "openai_results_shard*.jsonl"))):
        name = p.name
        if not name.startswith("openai_results_shard") or not name.endswith(".jsonl"):
            continue
        shard_s = name[len("openai_results_shard") : -len(".jsonl")]
        try:
            shard = int(shard_s)
        except ValueError:
            continue
        if want is not None and shard not in want:
            continue
        paths.append((shard, p))
    return paths


def _open_shard_files(
    out_dir: Path, shard_idx: int, *, want_openai: bool, want_gemini: bool
) -> tuple[IO[str] | None, IO[str] | None, IO[str]]:
    gemini_f = None
    openai_f = None
    if want_gemini:
        gemini_f = (out_dir / f"gemini_requests_shard{shard_idx:03d}.jsonl").open("w", encoding="utf-8")
    if want_openai:
        openai_f = (out_dir / f"openai_requests_shard{shard_idx:03d}.jsonl").open("w", encoding="utf-8")
    mapping_f = (out_dir / f"mapping_shard{shard_idx:03d}.jsonl").open("w", encoding="utf-8")
    return gemini_f, openai_f, mapping_f


def _close_files(files: Iterable[IO[str] | None]) -> None:
    for f in files:
        if f:
            f.close()


def _parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Export ordinance/amendment questionnaire batch requests (box-level).")
    ap.add_argument(
        "--classification-request-dir",
        required=True,
        help="Directory containing mapping_shard*.jsonl from the zoning box-classifier export",
    )
    ap.add_argument(
        "--classification-results-dir",
        required=True,
        help="Directory containing openai_results_shard*.jsonl from the zoning box-classifier download",
    )
    ap.add_argument("--output-dir", required=True, help="Directory to write request shards into")

    ap.add_argument(
        "--questions-xlsx",
        required=True,
        help="Path to Questions.xlsx",
    )
    ap.add_argument(
        "--questions-processed-sheet",
        default="Processed Info",
        help="Sheet name containing the processed question list (default: 'Processed Info')",
    )
    ap.add_argument(
        "--questions-input-info-sheet",
        default="Input Info",
        help="Sheet name containing question details / possible answers (default: 'Input Info')",
    )
    ap.add_argument(
        "--prompt-path",
        default="prompts/ordinance_box_questionnaire_prompt_text.txt",
        help="Prompt text file path (base instructions; questions/text are appended)",
    )

    ap.add_argument(
        "--provider",
        choices=["openai", "gemini", "both"],
        default="openai",
        help="Which provider request JSONLs to emit",
    )

    ap.add_argument(
        "--include-labels",
        default="full_ordinance,amendment_substantial,amendment_targeted",
        help="Comma-separated set of zoning labels to include from the classifier outputs",
    )
    ap.add_argument(
        "--use-present-flags",
        action="store_true",
        help="Also include boxes where classifier output has present.full_ordinance/amendment_* == true",
    )
    ap.add_argument(
        "--min-confidence",
        type=float,
        default=0.0,
        help="Minimum classifier confidence required to include a box (default: 0.0)",
    )

    ap.add_argument(
        "--shards",
        default=None,
        help="Optional shard selector for the classifier outputs (e.g. '0-16' or '0,1,2').",
    )
    ap.add_argument(
        "--allow-partial-results",
        action="store_true",
        help=(
            "Allow exporting from an incomplete set of classifier result shards. "
            "By default we fail if any mapping_shard is missing a corresponding openai_results_shard, "
            "to avoid silently generating partial downstream runs."
        ),
    )
    ap.add_argument("--max-requests", type=int, default=None, help="Stop after writing this many requests (smoke test)")
    ap.add_argument(
        "--exclude-from-request-dir",
        default=None,
        help=(
            "Optional request-dir from a prior questionnaire export. "
            "If provided, we will skip any custom_id already present in its mapping_shard*.jsonl "
            "(prevents duplicate re-submission across incremental exports)."
        ),
    )

    ap.add_argument(
        "--requests-per-shard",
        type=int,
        default=5000,
        help="Maximum requests per output shard file",
    )
    ap.add_argument(
        "--max-bytes-per-shard",
        type=int,
        default=0,
        help=(
            "Max bytes per provider request shard file (mapping is not capped). "
            "0 disables byte-based splitting. Recommended for OpenAI: ~180_000_000."
        ),
    )

    ap.add_argument("--openai-model", default="gpt-5-nano", help="OpenAI model name for /v1/responses batch body")
    ap.add_argument(
        "--openai-reasoning-effort",
        choices=["none", "minimal", "low", "medium", "high", "xhigh"],
        default="medium",
        help="OpenAI reasoning.effort for Responses API (set 'none' to omit)",
    )
    ap.add_argument(
        "--openai-max-output-tokens",
        type=int,
        default=None,
        help=(
            "OpenAI Responses API max_output_tokens. "
            "If omitted (recommended), we do not set it at all to avoid causing `status=incomplete`."
        ),
    )
    ap.add_argument(
        "--openai-text-format",
        choices=["json_object", "json_schema", "none"],
        default="none",
        help="OpenAI Responses text.format enforcement (default: none; we validate downstream).",
    )

    return ap.parse_args()


def main() -> None:
    args = _parse_args()

    cls_request_dir = Path(expanduser(args.classification_request_dir)).resolve()
    cls_results_dir = Path(expanduser(args.classification_results_dir)).resolve()
    out_dir = Path(expanduser(args.output_dir)).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    questions_xlsx = Path(expanduser(args.questions_xlsx)).resolve()
    prompt_path = Path(expanduser(args.prompt_path)).resolve()
    if not prompt_path.is_file():
        raise SystemExit(f"Prompt file not found: {prompt_path}")
    prompt_text = prompt_path.read_text(encoding="utf-8").strip()

    include_labels = {s.strip() for s in str(args.include_labels).split(",") if s.strip()}
    if not include_labels:
        raise SystemExit("--include-labels produced empty set")

    want_shards: set[int] | None = None
    if args.shards:
        want_shards = _parse_shard_selector(args.shards)
        if not want_shards:
            raise SystemExit(f"--shards parsed to empty set: {args.shards!r}")

    mapping_paths = _collect_mapping_paths(cls_request_dir, want=want_shards)
    if not mapping_paths:
        raise SystemExit(f"No mapping_shard*.jsonl found in {cls_request_dir}")

    result_paths = _collect_openai_result_paths(cls_results_dir, want=want_shards)
    if not result_paths:
        raise SystemExit(f"No openai_results_shard*.jsonl found in {cls_results_dir}")

    mapping_shards = {s for s, _ in mapping_paths}
    result_shards = {s for s, _ in result_paths}
    missing = sorted(mapping_shards - result_shards)
    if missing and not bool(args.allow_partial_results):
        preview = ",".join(str(s) for s in missing[:20])
        more = "" if len(missing) <= 20 else f" (+{len(missing) - 20} more)"
        raise SystemExit(
            "Classifier results are incomplete; refusing to export questionnaire requests.\n"
            f"Missing openai_results_shardNNN.jsonl for shards: {preview}{more}\n"
            "Run the downloader until all shards are present, OR re-run this exporter with either:\n"
            "  - --shards <subset> (export only completed shards), or\n"
            "  - --allow-partial-results (explicitly accept partial export)."
        )

    questions = _load_questions_xlsx(
        xlsx_path=questions_xlsx,
        processed_sheet=str(args.questions_processed_sheet),
        input_info_sheet=str(args.questions_input_info_sheet),
    )
    questions_payload = _questions_payload_for_prompt(questions)
    questions_json = json.dumps(questions_payload, ensure_ascii=False, separators=(",", ":"))
    question_set_sha = _sha256(json.dumps(questions_payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")))

    provider: Provider = args.provider  # type: ignore[assignment]
    want_openai = provider in {"openai", "both"}
    want_gemini = provider in {"gemini", "both"}

    exclude_ids: set[str] = set()
    if args.exclude_from_request_dir:
        prior_dir = Path(expanduser(str(args.exclude_from_request_dir))).resolve()
        if not prior_dir.is_dir():
            raise SystemExit(f"--exclude-from-request-dir is not a directory: {prior_dir}")
        for mp in sorted(prior_dir.glob("mapping_shard*.jsonl")):
            with mp.open("r", encoding="utf-8") as f:
                for raw in f:
                    line = raw.strip()
                    if not line:
                        continue
                    try:
                        obj = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    key = obj.get("id") or obj.get("custom_id")
                    if isinstance(key, str) and key:
                        exclude_ids.add(key)

    # Load mapping in memory so we can locate transcripts for each custom_id.
    mapping_by_id: dict[str, dict[str, Any]] = {}
    for shard, mp in mapping_paths:
        with mp.open("r", encoding="utf-8") as f:
            for raw in f:
                line = raw.strip()
                if not line:
                    continue
                obj = json.loads(line)
                key = obj.get("id") or obj.get("custom_id")
                if not isinstance(key, str) or not key:
                    continue
                mapping_by_id[key] = obj

    if not mapping_by_id:
        raise SystemExit(f"Loaded 0 mapping lines from {cls_request_dir}")

    # Prepare output shards.
    shard_idx = 0
    reqs_in_shard = 0
    bytes_in_shard_openai = 0
    bytes_in_shard_gemini = 0
    gemini_f, openai_f, mapping_f = _open_shard_files(out_dir, shard_idx, want_openai=want_openai, want_gemini=want_gemini)

    def rotate_if_needed(next_openai_line: str | None, next_gemini_line: str | None) -> None:
        nonlocal shard_idx, reqs_in_shard, bytes_in_shard_openai, bytes_in_shard_gemini, gemini_f, openai_f, mapping_f
        if reqs_in_shard <= 0:
            return
        if reqs_in_shard >= int(args.requests_per_shard):
            pass
        else:
            limit = int(args.max_bytes_per_shard)
            if limit > 0:
                if next_openai_line and (bytes_in_shard_openai + len(next_openai_line.encode("utf-8")) > limit):
                    pass
                elif next_gemini_line and (bytes_in_shard_gemini + len(next_gemini_line.encode("utf-8")) > limit):
                    pass
                else:
                    return
            else:
                return

        _close_files([gemini_f, openai_f, mapping_f])
        shard_idx += 1
        reqs_in_shard = 0
        bytes_in_shard_openai = 0
        bytes_in_shard_gemini = 0
        gemini_f, openai_f, mapping_f = _open_shard_files(out_dir, shard_idx, want_openai=want_openai, want_gemini=want_gemini)

    written = 0
    selected = 0
    skipped_excluded = 0
    skipped_no_mapping = 0
    skipped_not_selected = 0
    skipped_parse_error = 0
    skipped_not_completed = 0
    skipped_http_error = 0
    transcript_errors = 0
    label_counts = Counter()

    min_conf = float(args.min_confidence)

    def build_prompt(*, transcript: str) -> str:
        # Keep prompt deterministic and easy to diff/harden.
        return (
            f"{prompt_text}\n\n"
            f"QUESTIONS_JSON:\n{questions_json}\n\n"
            "OCR_TEXT:\n"
            f"{transcript}\n"
        )

    # Cache last loaded page to avoid re-parsing JSON repeatedly.
    last_page_path: Path | None = None
    last_page_boxes: dict[int, tuple[str, dict[str, Any]]] = {}

    def get_transcript_for(mapping_row: dict[str, Any]) -> tuple[str, dict[str, Any]]:
        nonlocal last_page_path, last_page_boxes
        page_path_raw = mapping_row.get("page_path")
        box_id = mapping_row.get("box_id")
        if not isinstance(page_path_raw, str) or not page_path_raw:
            raise ValueError("mapping missing page_path")
        if not isinstance(box_id, int):
            # Some older mapping files might store as str; coerce.
            try:
                box_id = int(box_id)
            except Exception as exc:
                raise ValueError(f"mapping missing/invalid box_id: {box_id!r}") from exc
        page_path = Path(expanduser(page_path_raw)).resolve()
        if not page_path.is_file():
            raise ValueError(f"page_path not found: {page_path}")

        if last_page_path != page_path:
            # Rebuild cache for the page.
            data = json.loads(page_path.read_text(encoding="utf-8"))
            boxes = data.get("boxes")
            if not isinstance(boxes, list):
                raise ValueError(f"Invalid page JSON (missing boxes list): {page_path}")
            last_page_boxes = {}
            for b in boxes:
                if not isinstance(b, dict):
                    continue
                bid = b.get("id")
                if not isinstance(bid, int):
                    continue
                status = b.get("status")
                transcript = (b.get("transcript") or "").strip()
                if status == "ok" and transcript:
                    stats = {
                        "source_model": data.get("model"),
                        "box_cls": b.get("class") or b.get("cls"),
                        "box_text_chars": len(transcript),
                        "box_text_sha256": _sha256(transcript),
                    }
                    last_page_boxes[bid] = (transcript, stats)
            last_page_path = page_path

        if box_id not in last_page_boxes:
            # Fall back to a targeted loader to produce a clearer error.
            return _load_box_transcript(page_path=page_path, box_id=int(box_id))
        return last_page_boxes[int(box_id)]

    def write_one(*, key: str, prompt: str, mapping_line: dict[str, Any]) -> None:
        nonlocal written, selected, reqs_in_shard, bytes_in_shard_openai, bytes_in_shard_gemini

        mapping_line_s = json.dumps(mapping_line, ensure_ascii=False) + "\n"

        gemini_line_s = None
        if gemini_f:
            gemini_line_s = (
                json.dumps(
                    {
                        "key": key,
                        "request": {
                            "contents": [
                                {"role": "user", "parts": [{"text": prompt}]},
                            ]
                        },
                    },
                    ensure_ascii=False,
                )
                + "\n"
            )

        openai_line_s = None
        if openai_f:
            body: dict[str, Any] = {
                "model": args.openai_model,
                "input": [{"role": "user", "content": [{"type": "input_text", "text": prompt}]}],
            }
            if args.openai_max_output_tokens is not None and int(args.openai_max_output_tokens) > 0:
                body["max_output_tokens"] = int(args.openai_max_output_tokens)
            text_cfg = _openai_text_config(args.openai_text_format)
            if text_cfg is not None:
                body["text"] = text_cfg
            if args.openai_reasoning_effort and args.openai_reasoning_effort != "none":
                body["reasoning"] = {"effort": args.openai_reasoning_effort}

            openai_line_s = (
                json.dumps(
                    {
                        "custom_id": key,
                        "method": "POST",
                        "url": "/v1/responses",
                        "body": body,
                    },
                    ensure_ascii=False,
                )
                + "\n"
            )

        rotate_if_needed(openai_line_s, gemini_line_s)

        mapping_f.write(mapping_line_s)
        if gemini_f and gemini_line_s:
            gemini_f.write(gemini_line_s)
            bytes_in_shard_gemini += len(gemini_line_s.encode("utf-8"))
        if openai_f and openai_line_s:
            openai_f.write(openai_line_s)
            bytes_in_shard_openai += len(openai_line_s.encode("utf-8"))

        reqs_in_shard += 1
        written += 1
        selected += 1

    # Process classifier results, selecting and exporting questionnaire requests.
    for shard, rp in result_paths:
        with rp.open("r", encoding="utf-8") as f:
            for raw in f:
                if args.max_requests is not None and written >= int(args.max_requests):
                    break

                line = raw.strip()
                if not line:
                    continue
                obj = json.loads(line)
                key = obj.get("custom_id")
                if not isinstance(key, str) or not key:
                    continue

                mapping_row = mapping_by_id.get(key)
                if mapping_row is None:
                    skipped_no_mapping += 1
                    continue

                resp = obj.get("response") or {}
                if not isinstance(resp, dict):
                    skipped_http_error += 1
                    continue

                status_code = resp.get("status_code")
                if status_code != 200:
                    skipped_http_error += 1
                    continue

                body = (resp.get("body") or {}) if isinstance(resp.get("body"), dict) else {}
                if not body:
                    skipped_parse_error += 1
                    continue

                if body.get("status") != "completed":
                    skipped_not_completed += 1
                    continue

                output_text = _extract_openai_output_text(body)
                if not output_text.strip():
                    skipped_parse_error += 1
                    continue

                try:
                    parsed = _coerce_json(output_text)
                except Exception:  # noqa: BLE001
                    skipped_parse_error += 1
                    continue

                label = parsed.get("label")
                conf = parsed.get("confidence")
                present = parsed.get("present")

                if isinstance(label, str):
                    label_counts[label] += 1

                try:
                    conf_f = float(conf) if isinstance(conf, (int, float)) else 0.0
                except Exception:
                    conf_f = 0.0

                include = False
                if isinstance(label, str) and label in include_labels and conf_f >= min_conf:
                    include = True
                elif args.use_present_flags and isinstance(present, dict) and conf_f >= min_conf:
                    for k in ["full_ordinance", "amendment_substantial", "amendment_targeted"]:
                        if present.get(k) is True and k in include_labels:
                            include = True
                            break

                if not include:
                    skipped_not_selected += 1
                    continue

                if exclude_ids and key in exclude_ids:
                    skipped_excluded += 1
                    continue

                try:
                    transcript, tstats = get_transcript_for(mapping_row)
                except Exception as exc:  # noqa: BLE001
                    transcript_errors += 1
                    _eprint(f"warn: failed loading transcript for {key}: {exc}")
                    continue

                prompt = build_prompt(transcript=transcript)

                out_mapping = {
                    "id": key,
                    "page_id": mapping_row.get("page_id"),
                    "box_id": mapping_row.get("box_id"),
                    "page_path": mapping_row.get("page_path"),
                    "cls": mapping_row.get("cls"),
                    "bbox": mapping_row.get("bbox"),
                    "source_model": tstats.get("source_model"),
                    "questionnaire_model_openai": args.openai_model if want_openai else None,
                    "questionnaire_prompt_path": str(prompt_path),
                    "questions_xlsx": str(questions_xlsx),
                    "questions_processed_sheet": str(args.questions_processed_sheet),
                    "questions_input_info_sheet": str(args.questions_input_info_sheet),
                    "question_count": len(questions),
                    "question_set_sha256": question_set_sha,
                    "classifier_label": label,
                    "classifier_confidence": conf_f,
                    "classifier_present": present if isinstance(present, dict) else None,
                    "box_text_chars": tstats.get("box_text_chars"),
                    "box_text_sha256": tstats.get("box_text_sha256"),
                }

                write_one(key=key, prompt=prompt, mapping_line=out_mapping)

        if args.max_requests is not None and written >= int(args.max_requests):
            break

    _close_files([gemini_f, openai_f, mapping_f])

    print(f"output_dir\t{out_dir}")
    print(f"question_count\t{len(questions)}")
    print(f"question_set_sha256\t{question_set_sha}")
    print(f"classifier_results_shards_seen\t{len(result_paths)}")
    print(f"classifier_mapping_keys_loaded\t{len(mapping_by_id)}")
    print(f"requests_written\t{written}")
    print(f"selected_boxes\t{selected}")
    print(f"skipped_excluded\t{skipped_excluded}")
    print(f"skipped_no_mapping\t{skipped_no_mapping}")
    print(f"skipped_not_selected\t{skipped_not_selected}")
    print(f"skipped_http_error\t{skipped_http_error}")
    print(f"skipped_not_completed\t{skipped_not_completed}")
    print(f"skipped_parse_error\t{skipped_parse_error}")
    print(f"transcript_errors\t{transcript_errors}")
    if label_counts:
        print("classifier_label_counts\t" + json.dumps(dict(label_counts), ensure_ascii=False))


if __name__ == "__main__":
    main()
